from flask import Flask, request
import requests
from openpyxl import Workbook, load_workbook
import os
import openpyxl
app = Flask(__name__)

# ===== Load tokens =====
def load_tokens(filename="popo.env"):
    tokens = {}
    with open(filename) as f:
        for line in f:
            line = line.strip()
            if line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            tokens[key] = value
    return tokens

tokens = load_tokens()
PAGE_ACCESS_TOKEN = tokens.get("PAGE_ACCESS_TOKEN")
VERIFY_TOKEN = tokens.get("VERIFY_TOKEN")
TELEGRAM_BOT_TOKEN = tokens.get("TELEGRAM_BOT_TOKEN")
TELEGRAM_CHAT_ID = tokens.get("TELEGRAM_CHAT_ID")
WHOLESALE_TELEGRAM_BOT_TOKEN = tokens.get("WHOLESALE_TELEGRAM_BOT_TOKEN")
WHOLESALE_TELEGRAM_CHAT_ID = tokens.get("WHOLESALE_TELEGRAM_CHAT_ID")
TRACKING_BOT_TOKEN = tokens.get("TRACKING_BOT_TOKEN")
TRACKING_CHAT_ID = tokens.get("TRACKING_CHAT_ID")

# ===== Products =====
PRODUCTS = {
    "خبز الشعير": 53,
    "خبز بذور الكتان": 54,
    "خبز الشوفان": 62,
    "خبز بذور الشيا": 62,
    "الخبز الاسمر": 54,
    "خبز عالي الألياف": 56,
    "خبز عالي البروتين": 69
}

# ===== Users =====
USER_ORDERS = {}
# في أعلى ملف الكود مع الإعدادات الأخرى
EXCEL_FILE = "orders.xlsx"  # استبدل orders.xlsx باسم ملفك الحقيقي

def ensure_excel_exists():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orders"
        # رؤوس الأعمدة بناءً على الصورة التي أرسلتها بالترتيب الصحيح
        headers = [
            "الاسم ثلاثي", "اسم المحافظة", "اسم المنطقة", 
            "اسم الشارع + علامة مميزة", "رقم العمارة", "رقم الشقة", 
            "رقم هاتف ويفضل يكون عليه واتساب", "رقم هاتف اخر (ان وجد)", 
            "الطلب", "الإجمالي بشحن", "التوصيل", "ه+A1:L2دية"
        ]
        ws.append(headers)
        wb.save(EXCEL_FILE)
        print(f"✅ تم إنشاء ملف جديد باسم {EXCEL_FILE}")

# استدعاء الدالة عند تشغيل البوت
ensure_excel_exists()
# ===== Bread Ingredients =====
BREAD_INGREDIENTS = {
    "خبز الشعير": "🟦 خبز الشعير → 53 سعر حراري\nدقيق الشعير\nدقيق القمح حبة كاملة\nخميرة طبيعية\nأملاح البحر بنسبة قليلة\nكل أنواع الخبز خالية من السكر والدهون واللبن والمواد الحافظة",
    "خبز الشوفان": "🟦 خبز الشوفان → 74 سعر حراري\nدقيق الشوفان\nدقيق القمح حبة كاملة\nخميرة طبيعية\nأملاح البحر بنسبة قليلة\nكل أنواع الخبز خالية من السكر والدهون واللبن والمواد الحافظة",
    "خبز الشيا": "🟦 خبز الشيا → 74 سعر حراري\nدقيق بذور الشيا\nدقيق القمح حبة كاملة\nخميرة طبيعية\nأملاح البحر بنسبة قليلة\nكل أنواع الخبز خالية من السكر والدهون واللبن والمواد الحافظة",
    "خبز الكتان": "🟪 خبز بذور الكتان → 77 سعر حراري\nدقيق بذور الكتان\nدقيق القمح حبة كاملة\nخميرة طبيعية\nأملاح البحر بنسبة قليلة\nكل أنواع الخبز خالية من السكر والدهون واللبن والمواد الحافظة",
    "الخبز الاسمر": "🟩 خبز أسمر → 70 سعر حراري\nدقيق القمح حبة كاملة\nنخالة القمح\nقليل من أملاح البحر والخميرة\nكل أنواع الخبز خالية من السكر والدهون واللبن والمواد الحافظة",
    "خبز عالي الألياف": "🟥 خبز عالي الألياف → 61 سعر حراري\nدقيق القمح حبة كاملة\nنخالة القمح\nقليل من أملاح البحر والخميرة\nكل أنواع الخبز خالية من السكر والدهون واللبن والمواد الحافظة",
    "خبز عالي البروتين": "🟧 خبز عالي البروتين → 58 سعر حراري\nبذور الكينوا\nدقيق جوز الهند\nدقيق اللوز\nدقيق القمح حبة كاملة\nقليل من أملاح البحر والخميرة\nكل أنواع الخبز خالية من السكر والدهون واللبن والمواد الحافظة"
}

def send_telegram_notification(message, bot_token=None, chat_id=None):
    # استخدام المتغيرات اللي أنت عرفتها في أول الكود كقيم افتراضية
    FINAL_TOKEN = bot_token if bot_token else TELEGRAM_BOT_TOKEN
    FINAL_CHAT_ID = chat_id if chat_id else TELEGRAM_CHAT_ID
    
    url = f"https://api.telegram.org/bot{FINAL_TOKEN}/sendMessage"
    payload = {
        "chat_id": FINAL_CHAT_ID,
        "text": message,
        "parse_mode": "Markdown"
    }
    
    try:
        import requests
        response = requests.post(url, json=payload)
        return response.json()
    except Exception as e:
        print(f"❌ فشل إرسال إشعار تليجرام: {e}")
def get_user_data_by_phone(phone_number):
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        # البحث من الأسفل للأعلى لجلب أحدث طلب للعميل
        for row in range(sheet.max_row, 1, -1):
            excel_phone = str(sheet.cell(row=row, column=7).value).strip() # عمود رقم الهاتف
            if excel_phone == str(phone_number).strip():
                return {
                    "الاسم ثلاثي": sheet.cell(row=row, column=1).value,
                    "اسم المحافظة": sheet.cell(row=row, column=2).value,
                    "اسم المنطقة": sheet.cell(row=row, column=3).value,
                    "اسم الشارع + علامة مميزة": sheet.cell(row=row, column=4).value,
                    "رقم العمارة": sheet.cell(row=row, column=5).value,
                    "رقم الشقة": sheet.cell(row=row, column=6).value,
                    "الطلب": sheet.cell(row=row, column=9).value,        # عمود I
                    "الإجمالي بشحن": sheet.cell(row=row, column=10).value, # عمود J
                    "التوصيل": sheet.cell(row=row, column=11).value      # عمود K
                }
    except Exception as e:
        print(f"Error reading excel: {e}")
    return None

def update_excel_row_directly(phone_number, combined_text, total_price, delivery_status, gift_text):
    try:
        import openpyxl
        wb = openpyxl.load_workbook("orders.xlsx")
        ws = wb.active
        
        target_row = None
        search_phone = str(phone_number).strip()
        
        # البحث عن رقم الهاتف في العمود رقم 7 (G)
        for row in range(ws.max_row, 1, -1):
            cell_value = str(ws.cell(row=row, column=7).value).strip()
            if cell_value == search_phone:
                target_row = row
                break
        
        if target_row:
            # تحديث الخانات بناءً على ترتيب أعمدة صورتك حصراً
            ws.cell(row=target_row, column=9).value = combined_text   # عمود الطلب (I)
            ws.cell(row=target_row, column=10).value = total_price    # عمود الإجمالي (J)
            ws.cell(row=target_row, column=11).value = delivery_status # عمود التوصيل (K)
            ws.cell(row=target_row, column=12).value = gift_text      # عمود الهدية (L)
            
            wb.save("orders.xlsx")
            return True
        return False
    except Exception as e:
        print(f"Error updating excel: {e}")
        return False
# ===== Save Order To Excel =====
import openpyxl

def save_order_to_excel(customer_data, order_text, total_price, delivery_text, gift_text):
    try:
        ensure_excel_exists()
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # الترتيب ده مطابق للصورة اللي انت بعتها (من A إلى L)
        new_row = [
            customer_data.get("الاسم ثلاثي", ""),                          # A - عمود 1
            customer_data.get("اسم المحافظة", ""),                        # B - عمود 2
            customer_data.get("اسم المنطقة", ""),                          # C - عمود 3
            customer_data.get("اسم الشارع + علامة مميزة", ""),             # D - عمود 4
            customer_data.get("رقم العمارة", ""),                          # E - عمود 5
            customer_data.get("رقم الشقة", ""),                           # F - عمود 6
            customer_data.get("رقم هاتف ويفضل يكون عليه واتساب", ""),      # G - عمود 7 (هنا المشكلة كانت)
            customer_data.get("رقم هاتف اخر (ان وجد)", ""),               # H - عمود 8
            order_text,                                                    # I - عمود 9 (الطلب)
            total_price,                                                   # J - عمود 10 (الإجمالي)
            delivery_text,                                                 # K - عمود 11 (التوصيل)
            gift_text                                                      # L - عمود 12 (الهدية)
        ]
        
        ws.append(new_row)
        wb.save(EXCEL_FILE)
        print("✅ تم الحفظ بنجاح في الأعمدة الصحيحة.")
    except Exception as e:
        print(f"❌ خطأ في الحفظ: {e}")
# --- دوال التعامل مع الإكسيل ---


def find_order_row_by_phone(phone_number):
    file_name = "orders.xlsx"
    if not os.path.exists(file_name): return None
    wb = load_workbook(file_name)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=False))
    # البحث من الأحدث (الأسفل) للأقدم لضمان الوصول لآخر أوردر للعميل
    for i in range(len(rows) - 1, 0, -1):
        if str(rows[i][6].value) == str(phone_number):
            return i + 1  # يرجع رقم السطر الفعلي في الإكسيل
    return None

def delete_order_from_excel(row_index):
    try:
        wb = load_workbook("orders.xlsx")
        ws = wb.active
        ws.delete_rows(row_index)
        wb.save("orders.xlsx")
        return True
    except Exception as e:
        print(f"Error deleting row: {e}")
        return False
    


def send_wholesale_telegram_notification(text):
    # استخدم التوكن الجديد الخاص ببوت الجملة
    url = f"https://api.telegram.org/bot{WHOLESALE_TELEGRAM_BOT_TOKEN}/sendMessage"
    payload = {"chat_id": WHOLESALE_TELEGRAM_CHAT_ID, "text": text}
    try:
        requests.post(url, json=payload)
    except Exception as e:
        print(f"Error sending wholesale telegram: {e}")


# ===== Wholesale Logic (معدلة للإرسال الفوري) =====
def save_wholesale_to_excel(data):
    file_name = "Wholesale.xlsx"
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.append(["الاسم", "المحافظة", "المنطقة", "محل أم أون لاين", "رقم التليفون"])
        wb.save(file_name)

    wb = load_workbook(file_name)
    ws = wb.active
    ws.append([data.get("الاسم"), data.get("المحافظة"), data.get("المنطقة"), data.get("محل أم أون لاين"), data.get("رقم التليفون")])
    wb.save(file_name)
    
    # إرسال الإشعار فوراً لبوت الجملة
    admin_msg = (
    "🏢 **طلب انضمام لعملاء الجملة**\n\n"
    f"👤 الاسم: {data.get('الاسم')}\n"
    f"📞 الهاتف: {data.get('رقم الهاتف')}\n"
    f"💼 النشاط: {data.get('نوع النشاط')}\n"
    f"📦 الكمية: {data.get('الكمية المطلوبة تقريبا')}"
    )
    send_telegram_notification(admin_msg)
####################################
def send_tracking_telegram_notification(text):
    print("--- محاولة إرسال إشعار للمتابعة ---") # عشان نتأكد إن الدالة اشتغلت أصلاً
    url = f"https://api.telegram.org/bot{TRACKING_BOT_TOKEN}/sendMessage"
    payload = {
        "chat_id": TRACKING_CHAT_ID, 
        "text": text
    }
    try:
        response = requests.post(url, json=payload)
        # السطر ده هو اللي هيقولنا "مبعتش ليه"
        print(f"نتيجة تليجرام: {response.status_code} - {response.text}")
    except Exception as e:
        print(f"❌ خطأ تقني في الاتصال: {e}")
####################################
def normalize_text(text):
    if not text: return ""
    text = text.strip().replace(" ", "") # حذف المسافات
    replacements = {
        "أ": "ا", "إ": "ا", "آ": "ا",
        "ة": "ه", "ى": "ي",
        "ال": "" # حذف ال التعريف للبحث المرن
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text

def get_distributors(city_input):
    city = normalize_text(city_input)
    
    # --- الفئة الأولى: محافظات التوصيل المباشر (بدون موزعين) ---
    if any(x in city for x in ["قاهره", "جيزه", "اكتوبر", "تجمع", "شروق"]):
        return "DIRECT_DELIVERY_ONLY" # علامة عشان نرد عليه برد التوصيل المباشر

    # --- الفئة الثانية: المحافظات اللي ليها موزعين (الداتا اللي بعتها) ---
    # المنصورة
    if any(x in city for x in ["منصوره", "دقهليه"]):
        return ("📍 المنصورة (الماركت - العنوان):\n"
                "🏪 قناة السويس - برج الميرلاند\n🏪 المختلط - ش فريدة حسان\n🏪 هايبر مارت - ش سعد زغلول\n🏪 الامام محمد عبده - امام بلاتوه\n🏪 المشاية السفلية - امام جزيرة الورد\n"
                "🏪 الترعة - ش الخلفاء الراشدين\n🏪 طلخا (1،2،3) - ش صلاح سالم والبحر الأعظم\n🏪 أجا - ش بورسعيد\n🏪 بلقاس - خلف بنك مصر\n🏪 شربين - بجوار المركز\n🏪 السنبلاوين - (برايم مارت، نص مشكل)")

    elif "دمياط" in city:
        if "جديده" in city:
            return ("📍 دمياط الجديدة:\n🏪 أسواق القوس (الدولي، المركزية، محبوب)\n🏪 أسواق الحياة - المركزية\n🏪 العمدة - 17\n🏪 أبو عمار - ش أبو الخير\n🏪 البوادي - الأولى\n🏪 السنبايطي - ش باب الحارة")
        else:
            return ("📍 دمياط القديمة:\n🏪 جوهرة هايبر - الشرباصي\n🏪 السيسي/تاج سحل - ش وزير\n🏪 أبو العينين - خلف المحطة\n🏪 الجيار - الشعبية\n🏪 الكانتو/فودة - ش نافع\n🏪 الزيدي - السنانية")

    elif any(x in city for x in ["اسكندريه", "عجمي", "بيطاش", "هانوفيل"]):
        # الإسكندرية فيها الحالتين، هنعرض الموزعين ونقوله متاح توصيل برضه
        return ("📍 الإسكندرية (الموزعين المعتمدين بالطلبات الخارجية):\n1️⃣ هيلثي العجمي\n2️⃣ بيت الجملة (الحديد والصلب)\n3️⃣ زهران (البيطاش)\n4️⃣ فتح الله (ستار، الهانوفيل، أبو يوسف)\n5️⃣ أبو الفضل (عين شمس، السماليهي)\n6️⃣ كارفور العروبة\n\n💡 علماً بأن خدمة التوصيل للمنازل متاحة أيضاً في الإسكندرية.")

    elif "اسماعيليه" in city:
        return "📍 الإسماعيلية:\n🏪 (العمدة، أهل الصفقة، نقاوة، بيتي وان، التعارف، ستار، سلمى، زياد، خديجة، عيد، الحياة، رينا، الحجاز، الفلسطيني، باندا، الغنيمي، العائلة، الدنيا بخير، الوفاء)"

    elif any(x in city for x in ["قليوبيه", "بنها", "طوخ", "قناطر", "منوفيه", "شبين", "سادات", "اشمون"]):
        return "📍 القليوبية والمنوفية:\n📞 للتواصل مع الموزع المعتمد: 01090468901"

    elif "محله" in city: return "📍 المحلة: هيبي سايد"
    elif "دسوق" in city or "كفر الشيخ" in city: return "📍 كفر الشيخ:\n🏪 دسوق (هيبي فود)\n🏪 كفر الشيخ (هيبي ميك)\n📞 خدمة التوصيل: 01113398933"
    elif "اسيوط" in city: return "📍 أسيوط: هيبي لايف / ثلاجة الحرمين"
    elif "بني سويف" in city: return "📍 بني سويف: بن سليمان"
    elif "بورسعيد" in city: return "📍 بورسعيد: أون سبورت"
    elif "ميت غمر" in city: return "📍 ميت غمر: الكانت"
    elif "شرم" in city: return "📍 شرم الشيخ: All In One Market"
    elif "سيناء" in city or "عريش" in city: return "📍 شمال سيناء:\n🏪 منفذ العريش (بجوار مسجد النصر)\n📞 التواصل: 01098949491 / 01221346226"
    elif "قنا" in city: return "📍 قنا:\n📞 أرقام خدمة التوصيل: 01553344300 / 01015401540"
    elif "بحيره" in city: return "📍 البحيرة:\n📞 رقم الموزع: 01558830006"

    # --- الفئة الثالثة: خارج النطاق ---
    return "OUT_OF_SCOPE"
# ===== Verify webhook =====
@app.route('/webhook', methods=['GET'])
def verify():
    mode = request.args.get("hub.mode")
    token = request.args.get("hub.verify_token")
    challenge = request.args.get("hub.challenge")
    if mode == "subscribe" and token == VERIFY_TOKEN:
        return challenge, 200
    return "Verification failed", 403

# ===== Webhook POST (تحديث لضمان وجود كل الحقول) =====
@app.route('/webhook', methods=['POST'])
def webhook():
    data = request.get_json()
    for entry in data.get("entry", []):
        for event in entry.get("messaging", []):
            sender_id = event["sender"]["id"]
            if sender_id not in USER_ORDERS:
                USER_ORDERS[sender_id] = {
                    "items": {},
                    "data_fields": [
                        "الاسم ثلاثي", "اسم المحافظة", "اسم المنطقة", 
                        "اسم الشارع + علامة مميزة", "رقم العمارة", "رقم الشقة", 
                        "رقم هاتف ويفضل يكون عليه واتساب", "رقم هاتف اخر (ان وجد)"
                    ],
                    "wholesale_fields": ["الاسم", "المحافظة", "المنطقة", "محل أم أون لاين", "رقم التليفون"],
                    "current_question": 0,
                    "current_wholesale_question": 0,
                    "customer_data": {},
                    "wholesale_data": {},
                    "stage": "welcome"
                }

            if "message" in event and not event["message"].get("is_echo", False):
                if "quick_reply" in event["message"]:
                    payload = event["message"]["quick_reply"]["payload"]
                    handle_postback(sender_id, {"payload": payload})
                else:
                    handle_message(sender_id, event["message"])
            elif "postback" in event:
                handle_postback(sender_id, event["postback"])
    return "ok", 200
# ===== Send Messages =====
def send_message(recipient_id, text):
    url = f"https://graph.facebook.com/v16.0/me/messages?access_token={PAGE_ACCESS_TOKEN}"
    payload = {"recipient": {"id": recipient_id}, "message": {"text": text}}
    requests.post(url, json=payload)

def send_quick_replies(recipient_id, text, quick_replies):
    url = f"https://graph.facebook.com/v16.0/me/messages?access_token={PAGE_ACCESS_TOKEN}"
    payload = {"recipient": {"id": recipient_id}, "message": {"text": text, "quick_replies": quick_replies}}
    requests.post(url, json=payload)

# ===== Main Menu =====
def send_main_menu(sender_id):
    quick_replies = [
        {"content_type": "text", "title": "ℹ️ استفسار عن منتج", "payload": "INQUIRY_MENU"},
        {"content_type": "text", "title": "🛒 طلب أوردر", "payload": "START_ORDER"},
        {"content_type": "text", "title": "📦 متابعة/تعديل طلبك", "payload": "TRACK_ORDER_MENU"},
        {"content_type": "text", "title": "📍 أماكن توافرنا", "payload": "FIND_DISTRIBUTORS"}, # الزر الجديد
        {"content_type": "text", "title": "🏢 طلبات الجملة", "payload": "INQ_WHOLESALE"}
    ]
    send_quick_replies(sender_id, "مرحباً بك في خبز ريف 💚\nاختر أحد الخيارات:", quick_replies)
# ===== Welcome Message =====
def send_welcome(sender_id):
    text = (
        "شكراً لتواصلك مع خبز ريف 💚\n\n"
        "🎉 عرض رمضان:\n"
        "عند طلب 5 أكياس ➜ التوصيل مجاني 🚚\n"
        "عند طلب 8 أكياس ➜ التوصيل مجاني + كيس هدية 🎁"
    )
    quick_replies = [
        {"content_type": "text", "title": "ℹ️ استفسار عن منتج", "payload": "INQUIRY_MENU"},
        {"content_type": "text", "title": "🛒 طلب أوردر", "payload": "START_ORDER"},
        {"content_type": "text", "title": "📦 متابعة/تعديل طلبك", "payload": "TRACK_ORDER_MENU"},
        {"content_type": "text", "title": "📍 أماكن توافرنا", "payload": "FIND_DISTRIBUTORS"}, # الزر الجديد
        {"content_type": "text", "title": "🏢 طلبات الجملة", "payload": "INQ_WHOLESALE"}
    ]
    send_quick_replies(sender_id, text, quick_replies)

# ===== handle_message (تعديلات الاستقرار) =====
def handle_message(sender_id, message):
    user = USER_ORDERS.get(sender_id)
    text = message.get("text", "").strip()
    if not text: return

    # 1. جمع بيانات الأوردر العادي
    # 1. جمع بيانات الأوردر العادي
    if user["stage"] == "collecting_data":
        field = user["data_fields"][user["current_question"]]
        
        # فحص رقم الهاتف
        if field == "رقم هاتف ويفضل يكون عليه واتساب":
            if not (text.isdigit() and len(text) == 11):
                send_message(sender_id, "🚫 رقم غير صحيح! ارسل رقم صحيح مكون من 11 رقم.")
                return
            
            existing_data = get_user_data_by_phone(text)
            if existing_data:
                user["customer_data"] = existing_data
                user["stage"] = "confirm_existing_data"
                summary = (
                    f"👋 أهلاً بك من جديد يا {existing_data.get('الاسم ثلاثي', 'عميلنا العزيز')}!\n"
                    f"📍 العنوان المسجل: {existing_data.get('اسم المحافظة')} - {existing_data.get('اسم المنطقة')}\n"
                    "هل تريد استخدام نفس البيانات السابقة؟"
                )
                quick_replies = [
                    {"content_type": "text", "title": "✅ نعم، استخدمها", "payload": "USE_OLD_DATA"},
                    {"content_type": "text", "title": "✏️ لا، بيانات جديدة", "payload": "RE-ENTER_DATA"}
                ]
                send_quick_replies(sender_id, summary, quick_replies)
                return

        # فحص المحافظة
        if field == "اسم المحافظة":
            allowed = ["القاهرة","قاهره","قاهرة","القاهره","الجيزة","الجيزه","الاسكندرية","الاسكندريه","الإسكندرية","إسكندرية","القليوبية","قليوبية","قليوبيه","القليوبيه"]
            if text not in allowed:
                send_message(sender_id, "❌ نأسف 🙏 المحافظة خارج نطاق التوصيل المباشر حالياً.")
                user["stage"] = "welcome"
                send_main_menu(sender_id)
                return

        # فحص خاص بالقليوبية عند سؤال المنطقة
        if field == "اسم المنطقة":
            gov = user["customer_data"].get("اسم المحافظة", "")
            allowed_qalyubia = ["العبور", "شبرا الخيمة", "شبرا الخيمه", "الخصوص"]
            
            if "قليوبية" in gov or "القليوبية" in gov:
                if text not in allowed_qalyubia:
                    msg = (
                        f"عذراً، منطقة '{text}' في القليوبية متاح لها موزعين فقط حالياً. 😔\n\n"
                        "المناطق المتاحة للتوصيل المباشر: (العبور - شبرا الخيمة - الخصوص).\n"
                        "يمكنك البحث عن أقرب موزع لك من القائمة الرئيسية."
                    )
                    send_message(sender_id, msg)
                    user["stage"] = "welcome"
                    send_main_menu(sender_id)
                    return

        # حفظ البيانات الحالية والانتقال للسؤال التالي
        user["customer_data"][field] = text
        user["current_question"] += 1
        
        if user["current_question"] < len(user["data_fields"]):
            ask_next_question(sender_id)
        else:
            user["stage"] = "choosing_products"
            send_products(sender_id)
        return

    # 2. البحث عن الموزعين
    elif user["stage"] == "search_distributor":
        result = get_distributors(text)
        if result == "DIRECT_DELIVERY_ONLY":
            send_message(sender_id, "📍 هذه المنطقة متاح بها توصيل للمنازل فقط.\nاضغط '🛒 طلب أوردر' للبدء.")
        elif result == "OUT_OF_SCOPE":
            send_message(sender_id, "بعتذر لحضرتك ولكن منطقة حضرتك خارج حيز التوصيل حالياً. 😔")
        else:
            send_message(sender_id, result)
        
        user["stage"] = "welcome"
        send_main_menu(sender_id)
        return

# 3. بيانات الجملة (Wholesale)
    elif user["stage"] == "wholesale":
        fields = user.get("wholesale_fields", [])
        current_idx = user.get("current_wholesale_question", 0)

        # شرط الأمان: التأكد أن الاندكس داخل نطاق القائمة
        if current_idx < len(fields):
            user["wholesale_data"][fields[current_idx]] = text
            user["current_wholesale_question"] += 1
            
            # فحص بعد الزيادة: هل لسه فيه أسئلة تانية؟
            if user["current_wholesale_question"] < len(fields):
                next_q = fields[user["current_wholesale_question"]]
                send_message(sender_id, f"من فضلك اكتب {next_q}:")
            else:
                # خلصنا كل الأسئلة
                save_wholesale_to_excel(user["wholesale_data"])
                send_message(sender_id, "✅ تم تسجيل بياناتك بنجاح. سيتواصل معك قسم الجملة قريباً. 💚")
                # إعادة تصفير البيانات
                user["stage"] = "welcome"
                user["current_wholesale_question"] = 0
                user["wholesale_data"] = {} 
                send_main_menu(sender_id)
        else:
            # في حالة حدوث خطأ والوصول لاندكس خارج النطاق
            user["stage"] = "welcome"
            user["current_wholesale_question"] = 0
            send_main_menu(sender_id)
        return
# 4. البحث عن رقم الهاتف للمتابعة/التعديل
    elif user["stage"] == "track_ask_phone":
        # ... (كود فحص رقم الهاتف) ...
        existing_data = get_user_data_by_phone(text)
        
        if existing_data:
            # 1. أولاً: استخراج تفاصيل الأوردر من البيانات
            # تأكد أن كلمة 'الطلب' هي اسم العمود الصحيح عندك في الإكسيل
            last_order_details = existing_data.get('الطلب', 'لا يوجد طلبات سابقة')
            
            user["customer_data"] = existing_data
            user["temp_phone"] = text
            user["stage"] = "order_found_options"
            
            # 2. ثانياً: وضع المتغير داخل الرسالة
            summary = (
                f"✅ تم العثور على بياناتك يا {existing_data.get('الاسم ثلاثي', 'فندم')}!\n"
                f"📍 العنوان: {existing_data.get('اسم المحافظة')} - {existing_data.get('اسم المنطقة')}\n"
                f"📦 أخر أوردر ليك كان: ({last_order_details})\n\n"
                "كيف يمكننا مساعدتك اليوم؟"
            )
            quick_replies = [
                {"content_type": "text", "title": "🔍 استفسار عن الحالة", "payload": "TRACK_INQUIRY"},
                {"content_type": "text", "title": "➕ إضافة أصناف", "payload": "MODIFY_ORDER_MENU"},
                {"content_type": "text", "title": "❌ إلغاء الطلب", "payload": "CANCEL_EXISTING_ORDER"},
                {"content_type": "text", "title": "🏠 القائمة الرئيسية", "payload": "MAIN_MENU"}
            ]
            send_quick_replies(sender_id, summary, quick_replies)
        
        else:
            # لو الرقم مش موجود (الطريقة الشيك)
            msg = (
                "لم نجد أي طلبات مسجلة بهذا الرقم حالياً 🧐\n\n"
                "ربما تم كتابة الرقم بشكل خاطئ؟ أو ربما لم تجرب طعم 'خبز ريف' حتى الآن! 💚✨\n"
                "يسعدنا جداً أن تنضم إلينا وتطلب أوردرك الأول الآن."
            )
            quick_replies = [
                {"content_type": "text", "title": "🛒 اطلب أوردر جديد", "payload": "START_ORDER"},
                {"content_type": "text", "title": "🔢 تجربة رقم آخر", "payload": "TRACK_ORDER_MENU"},
                {"content_type": "text", "title": "🏠 العودة للرئيسية", "payload": "MAIN_MENU"}
            ]
            send_quick_replies(sender_id, msg, quick_replies)
        return
    # العودة للقائمة الرئيسية لو تاه
    if user["stage"] == "welcome" or text.lower() in ["menu", "القائمة", "الرئيسية"]:
        send_welcome(sender_id)
        
def send_inquiry_options(sender_id):
    quick_replies = [
        {"content_type": "text", "title": "1️⃣ الأسعار", "payload": "INQ_PRICES"},
        {"content_type": "text", "title": "2️⃣ العروض المتاحة", "payload": "INQ_OFFERS"},
        {"content_type": "text", "title": "3️⃣ مكونات الخبز", "payload": "INQ_INGREDIENTS"},
        {"content_type": "text", "title": "4️⃣ كيفية حفظ المنتج", "payload": "INQ_STORAGE"},
        {"content_type": "text", "title": "5️⃣ الجلوتين", "payload": "INQ_GLUTEN"},
        {"content_type": "text", "title": "🏠 العودة للقائمة الرئيسية", "payload": "MAIN_MENU"}
    ]
    send_quick_replies(sender_id, "اختر نوع الاستفسار:", quick_replies)

def handle_inquiry(sender_id, payload):

    # =============================
    # الأسعار
    # =============================
    if payload == "INQ_PRICES":
        text = (
            "💰 أسعار الخبز:\n"
            "خبز الشعير: 53 جنيه\n"
            "خبز الشوفان: 62 جنيه\n"
            "خبز الشيا: 62 جنيه\n"
            "خبز الكتان: 54 جنيه\n"
            "خبز أسمر: 54 جنيه\n"
            "خبز عالي الألياف: 56 جنيه\n"
            "خبز عالي البروتين: 69 جنيه"
        )
        send_message(sender_id, text)

    # =============================
    # العروض
    # =============================
    elif payload == "INQ_OFFERS":
        text = (
            "🎉 عروض رمضان المبارك:\n"
            "✅ التوصيل مجاني عند طلب 5 أكياس\n"
            "✅ كيس هدية + توصيل مجاني عند طلب 8 أكياس"
        )
        send_message(sender_id, text)

    # =============================
    # مكونات الخبز (يعرض الأنواع فقط)
    # =============================
    elif payload == "INQ_INGREDIENTS":

        quick_replies = []

        for bread_name in BREAD_INGREDIENTS.keys():
            quick_replies.append({
                "content_type": "text",
                "title": bread_name,
                "payload": f"ING_{bread_name}"
            })

        quick_replies.append({
            "content_type": "text",
            "title": "🏠 القائمة الرئيسية",
            "payload": "MAIN_MENU"
        })

        send_quick_replies(sender_id, "اختر نوع الخبز لعرض المكونات 👇", quick_replies)
        return

    # =============================
    # التخزين
    # =============================
    elif payload == "INQ_STORAGE":
        text = (
            "📦 كيفية حفظ المنتج:\n"
            "❄️ يخزن في الفريزر لمدة 6 أشهر\n"
            "🧊 يخزن في الثلاجة لمدة شهر\n"
            "🌡️ يخزن خارج الثلاجة لمدة 10 أيام\n"
            "⏳ بعد الخروج من الفريزر/الثلاجة اتركه دقائق للفك بدون تسخين"
        )
        send_message(sender_id, text)

    # =============================
    # الجلوتين
    # =============================
    elif payload == "INQ_GLUTEN":
        text = (
        "أهم المعلومات الصحية عن خبز ريف 🌱\n"
        "- جميع الأنواع خالية تمامًا من الدقيق الأبيض\n"
        "- تحتوي كل الأنواع على نسبة قليلة جداً من الدقيق الأسمر لا تتعدّى 15% للباكيت\n"
        "- كما تحتوي على نسبة منخفضة جداً من الجلوتين لا تتعدّى 15% للباكيت\n\n"
        "علشان كدا خبز ريف اختيار صحي ومتوازن، ومناسب لأنظمة غذائية مختلفة 💚"
        )
        send_message(sender_id, text)

# =============================
    # الجملة
    # =============================
    elif payload == "INQ_WHOLESALE":
        USER_ORDERS[sender_id]["stage"] = "wholesale"
        USER_ORDERS[sender_id]["wholesale_data"] = {}
        USER_ORDERS[sender_id]["wholesale_fields"] = [
            "الاسم",
            "المحافظة",
            "المنطقة",
            "رقم التليفون"
        ]
        USER_ORDERS[sender_id]["current_wholesale_question"] = 0
        send_message(sender_id, "من فضلك اكتب الاسم:") 
        return # ضروري جداً لمنع الكود من إكمال إرسال أزرار الاستفسارات

    elif payload == "MAIN_MENU":
        send_main_menu(sender_id)
        return

    # =============================
    # أزرار المتابعة بعد أي استفسار
    # =============================
    quick_replies = [
        {"content_type": "text", "title": "🛒 طلب أوردر", "payload": "START_ORDER"},
        {"content_type": "text", "title": "🔙 القائمة السابقة", "payload": "INQUIRY_MENU"},
        {"content_type": "text", "title": "🏠 القائمة الرئيسية", "payload": "MAIN_MENU"}
    ]

    send_quick_replies(sender_id, "اختر أحد الخيارات:", quick_replies)

# ... (الدوال السابقة مثل send_message و save_order_to_excel)

# 1. ضع الدالة هنا
def process_order_action(sender_id, action_type):
    user = USER_ORDERS.get(sender_id, {})
    # جلب البيانات التي تم سحبها من الإكسيل في الخطوة السابقة
    data = user.get("customer_data", {})
    
    # 1. تجهيز نص الرسالة الموحد بالتنسيق التفصيلي
    # استخدمنا الرموز التعبيرية (Emojis) لتمييز نوع الطلب (إلغاء ⚠️ أو استفسار ❓)
    header_emoji = "⚠️" if action_type == "إلغاء" else "❓"
    
    admin_msg = (
        f"{header_emoji} **طلب {action_type} أوردر قائم!**\n\n"
        "👤 **بيانات العميل:**\n"
        f"الاسم ثلاثي: {data.get('الاسم ثلاثي', '')}\n"
        f"اسم المحافظة: {data.get('اسم المحافظة', '')}\n"
        f"اسم المنطقة: {data.get('اسم المنطقة', '')}\n"
        f"اسم الشارع + علامة مميزة: {data.get('اسم الشارع + علامة مميزة', '')}\n"
        f"رقم العمارة: {data.get('رقم العمارة', '')}\n"
        f"رقم الشقة: {data.get('رقم الشقة', '')}\n"
        f"رقم هاتف ويفضل يكون عليه واتساب: {data.get('رقم هاتف ويفضل يكون عليه واتساب', '')}\n"
        f"رقم هاتف اخر (ان وجد): {data.get('رقم هاتف اخر (ان وجد)', '')}\n\n"
        "📦 **تفاصيل الأوردر الأخير:**\n"
        f"{data.get('الطلب', 'غير متوفر')}\n"
        "-----------------\n"
        f"🛠️ **نوع الإجراء المطلوب:** {action_type}"
    )

    # 2. إرسال الإشعار لجروب المتابعة (Tracking)
    # نستخدم المتغيرات العالمية TRACKING_BOT_TOKEN و TRACKING_CHAT_ID
    send_telegram_notification(
        admin_msg, 
        TRACKING_BOT_TOKEN, 
        TRACKING_CHAT_ID
    )

    # 3. الرد على العميل في فيسبوك بناءً على نوع الطلب
    if action_type == "إلغاء":
        response_text = "✅ تم إرسال طلب الإلغاء للإدارة، وسيتم التأكيد معك قريباً. 💚"
    else:
        response_text = "✅ تم إرسال استفسارك لقسم المتابعة، وسيتم الرد عليك فوراً. 💚"
    
    send_message(sender_id, response_text)

    # 4. إعادة العميل للقائمة الرئيسية وتصفير الـ stage
    user["stage"] = "welcome"
    send_main_menu(sender_id)
# 2. ثم تليها دالة handle_postback التي تستخدمها

def handle_postback(sender_id, postback):
    payload = postback.get("payload")
    user = USER_ORDERS[sender_id]

    # =============================
    # بدء أوردر جديد
    # =============================
    if payload == "START_ORDER":
        USER_ORDERS[sender_id] = {
            "items": {},
            "data_fields": [
                "رقم هاتف ويفضل يكون عليه واتساب", 
                "الاسم ثلاثي",
                "اسم المحافظة",
                "اسم المنطقة",
                "اسم الشارع + علامة مميزة",
                "رقم العمارة",
                "رقم الشقة",
                "رقم هاتف اخر (ان وجد)"
            ],
            "current_question": 0,
            "customer_data": {},
            "stage": "collecting_data"
        }
        ask_next_question(sender_id)

    # =============================
    # التعامل مع البيانات القديمة
    # =============================
    elif payload == "USE_OLD_DATA":
        user["stage"] = "ordering"
        send_products(sender_id)

    elif payload == "RE-ENTER_DATA":
        user["current_question"] = 1 
        user["stage"] = "collecting_data"
        ask_next_question(sender_id)

    # =============================
    # قائمة المتابعة والتعديل
    # =============================
    elif payload == "TRACK_ORDER_MENU":
        user["stage"] = "track_ask_phone"
        send_message(sender_id, "من فضلك أدخل رقم الهاتف الذي قمت بعمل الطلب به (11 رقم):")

    elif payload == "TRACK_INQUIRY":
        user["stage"] = "processing_track_inquiry"
        process_order_action(sender_id, "استفسار")

    elif payload == "MODIFY_ORDER_MENU":
        quick_replies = [
            {"content_type": "text", "title": "➕ إضافة منتج", "payload": "ADD_TO_EXISTING"},
            {"content_type": "text", "title": "🔄 تغيير الأوردر بالكامل", "payload": "CHANGE_ENTIRE_ORDER"}
        ]
        send_quick_replies(sender_id, "هل تود إضافة منتج جديد للطلب أم تغيير الأصناف الحالية؟", quick_replies)

    # إضافة منتج لطلب قديم
    elif payload == "ADD_TO_EXISTING":
        user["stage"] = "adding_to_existing"
        user["items"] = {}  # تصفير السلة لاستقبال الإضافات الجديدة فقط
        send_message(sender_id, "قائمة الإضافات المتاحة 👇")
        send_products(sender_id)

    elif payload == "CHANGE_ENTIRE_ORDER":
        user["stage"] = "ordering"
        user["items"] = {}
        send_products(sender_id)

    elif payload == "CANCEL_EXISTING_ORDER":
        process_order_action(sender_id, "إلغاء")

    # =============================
    # الاستفسارات والمكونات
    # =============================
    elif payload == "INQUIRY_MENU":
        send_inquiry_options(sender_id)

    elif payload.startswith("INQ_"):
        handle_inquiry(sender_id, payload)

    elif payload.startswith("ING_"):
        bread_name = payload.replace("ING_", "")
        if bread_name in BREAD_INGREDIENTS:
            send_message(sender_id, BREAD_INGREDIENTS[bread_name])
            quick_replies = [
                {"content_type": "text", "title": "🛒 طلب أوردر", "payload": "START_ORDER"},
                {"content_type": "text", "title": "🔙 القائمة السابقة", "payload": "INQUIRY_MENU"},
                {"content_type": "text", "title": "🏠 القائمة الرئيسية", "payload": "MAIN_MENU"}
            ]
            send_quick_replies(sender_id, "اختر أحد الخيارات:", quick_replies)

    # =============================
    # عملية اختيار المنتجات
    # =============================
    elif payload.startswith("PRODUCT_"):
        product = payload.replace("PRODUCT_", "")
        user["selected_product"] = product
        send_quantity_menu(sender_id, product)

    elif payload.startswith("QTY_"):
        qty = int(payload.split("_")[1])
        product = user["selected_product"]
        user["items"][product] = user["items"].get(product, 0) + qty
        send_after_product_menu(sender_id)

    # =============================
    # إنهاء وتأكيد
    # =============================
    elif payload == "ADD_MORE":
        send_products(sender_id)

    elif payload == "FINISH_ORDER":
        show_final_summary(sender_id)

    elif payload == "CONFIRM_ORDER":
        confirm_order(sender_id)

    elif payload == "CANCEL_ORDER":
        cancel_order(sender_id)

    # =============================
    # موزعين وقائمة رئيسية
    # =============================
    elif payload == "FIND_DISTRIBUTORS":
        user["stage"] = "search_distributor"
        send_message(sender_id, "من فضلك اكتب اسم المحافظة للبحث عن أقرب موزع لك:")

    elif payload == "MAIN_MENU":
        send_main_menu(sender_id)

        
# ===== أسئلة جمع البيانات =====
def ask_next_question(sender_id):
    user = USER_ORDERS[sender_id]
    index = user["current_question"]
    if index < len(user["data_fields"]):
        field = user["data_fields"][index]
        send_message(sender_id, f"من فضلك اكتب {field}:")
    else:
        user["stage"] = "ordering"
        send_products(sender_id)

# ===== عرض المنتجات =====
def send_products(sender_id):
    quick_replies = [
        {"content_type": "text", "title": f"{name} - {price}ج", "payload": f"PRODUCT_{name}"}
        for name, price in PRODUCTS.items()
    ]
    send_quick_replies(sender_id, "اختر المنتج:", quick_replies)

def send_quantity_menu(sender_id, product):
    quick_replies = [
        {"content_type": "text", "title": str(i), "payload": f"QTY_{i}"}
        for i in range(1, 11)
    ]
    send_quick_replies(sender_id, f"كم عدد أكياس {product}؟", quick_replies)

def send_after_product_menu(sender_id):
    quick_replies = [
        {"content_type": "text", "title": "➕ طلب منتج اخر", "payload": "ADD_MORE"},
        {"content_type": "text", "title": "✅ إنهاء الأوردر", "payload": "FINISH_ORDER"}
    ]
    send_quick_replies(sender_id, "تم إضافة المنتج 👌", quick_replies)
import re

def extract_total_qty_from_text(order_text):
    # يبحث عن الأرقام التي تلي حرف x مثل x2 أو x10
    quantities = re.findall(r'x(\d+)', order_text)
    return sum(int(q) for q in quantities)

    
import re

import re

def show_final_summary(sender_id):
    user = USER_ORDERS.get(sender_id)
    if not user: return
    
    order = user.get("items", {})
    if not order:
        send_message(sender_id, "لم يتم اختيار أي منتجات بعد.")
        return

    # --- 1. حسابات الأصناف الجديدة (دائماً موجودة) ---
    new_items_qty = sum(order.values())
    new_items_price = sum(PRODUCTS[name] * qty for name, qty in order.items())
    details = "\n".join([f"✨ {name} x{qty} = {PRODUCTS[name]*qty}ج" for name, qty in order.items()])

    # --- 2. تعريف متغيرات الحالة السابقة بقيم افتراضية (لمنع الـ NameError) ---
    old_products_price = 0
    old_order_text = ""
    total_combined_qty = new_items_qty
    delivery_status = ""

    # --- 3. حالة إضافة على طلب قائم ---
    if user.get("stage") == "adding_to_existing":
        data = user.get("customer_data", {})
        old_order_text = str(data.get('الطلب') or data.get('الأوردر') or '')
        
        # البحث المرن عن السعر القديم
        raw_old_total = "0"
        for key, value in data.items():
            if "الإجمالي" in str(key):
                raw_old_total = str(value)
                break
        
        # تنظيف وتحويل السعر
        clean_old_total = re.sub(r'[^\d.]', '', raw_old_total)
        try:
            old_total_val = float(clean_old_total) if clean_old_total else 0.0
        except:
            old_total_val = 0.0

        # حساب الكمية القديمة
        old_qty_list = re.findall(r'x(\d+)', old_order_text)
        old_qty = sum(int(q) for q in old_qty_list)

        # حساب الشحن القديم لخصمه (لو الكمية < 5 كان فيه 30ج شحن)
        old_delivery = 30 if (old_qty > 0 and old_qty < 5) else 0
        old_products_price = old_total_val - old_delivery

        # الحسابات الكلية
        total_combined_qty = old_qty + new_items_qty
        total_products_price = old_products_price + new_items_price
        
        new_delivery = 0 if total_combined_qty >= 5 else 30
        final_grand_total = total_products_price + new_delivery

        if old_delivery == 30 and new_delivery == 0:
            delivery_status = "🚚 التوصيل: **مجاني** (بدلاً من 30ج) 🎉"
        elif new_delivery == 0:
            delivery_status = "🚚 التوصيل: **مجاني** ✨"
        else:
            delivery_status = f"🚚 التوصيل: {new_delivery}ج"

        summary = (
            "🧾 **ملخص تحديث الطلب:**\n\n"
            "📦 **الطلب السابق:**\n"
            f"{old_order_text}\n"
            f"💰 قيمة المنتجات السابقة: {old_products_price}ج\n"
            "-----------------\n"
            "➕ **الإضافات الجديدة:**\n"
            f"{details}\n"
            f"💵 قيمة الإضافات: {new_items_price}ج\n"
            "-----------------\n"
            f"📊 إجمالي الكمية: {total_combined_qty} أكياس\n"
            f"{delivery_status}\n"
            f"✅ **الإجمالي النهائي الجديد: {final_grand_total}ج**\n\n"
            "💡 تم دمج طلباتك وتحديث مصاريف الشحن تلقائياً."
        )
    
    # --- 4. حالة أوردر جديد تماماً ---
    else:
        delivery = 0 if new_items_qty >= 5 else 30
        total_price = new_items_price + delivery
        delivery_text = "مجاني" if delivery == 0 else f"{delivery}ج"
        
        summary = (
            "🧾 **ملخص طلبك:**\n\n"
            f"{details}\n"
            "-----------------\n"
            f"المجموع: {new_items_price}ج\n"
            f"التوصيل: {delivery_text}\n"
            f"الإجمالي: {total_price}ج"
        )

    quick_replies = [
        {"content_type": "text", "title": "✅ تأكيد وإرسال", "payload": "CONFIRM_ORDER"},
        {"content_type": "text", "title": "❌ إلغاء", "payload": "CANCEL_ORDER"}
    ]
    send_quick_replies(sender_id, summary, quick_replies)
# === أضف الدالة هنا ===
def update_existing_order_with_new_items(sender_id):
    user = USER_ORDERS[sender_id]
    data = user.get("customer_data", {})
    
    # 1. سحب البيانات القديمة (نص الطلب)
    old_order_text = data.get('الطلب', '')

    # 2. حساب عدد الأكياس القديمة وسعرها (بدون شحن)
    # سنحتاج أن نكون مخزنين "إجمالي المنتجات" و "إجمالي الكمية" في الإكسيل
    # لو مش مخزنينهم، هنحاول نسحبهم من السعر الإجمالي المسجل
    try:
        # بنسحب السعر القديم (ونفترض إنه كان شامل شحن، فهنشيل منه الـ 30 لو كان أقل من 5 أكياس)
        # الأفضل مستقبلاً نخزن خانة "سعر المنتجات" منفصلة في الإكسيل
        raw_total = str(data.get('الإجمالي بشحن', '0')).replace('ج', '').strip()
        old_total_with_shipping = float(raw_total)
        
        # بنعرف كان عليه شحن ولا لا من نص "هدية" أو "توصيل" في الإكسيل
        old_delivery = 0 if "مجاني" in str(data.get('التوصيل', '')) else 30
        old_products_price = old_total_with_shipping - old_delivery
    except:
        old_products_price = 0.0

    # 3. حساب الإضافات الجديدة
    new_items = user["items"]
    new_qty = sum(new_items.values())
    new_price = sum(PRODUCTS[name] * qty for name, qty in new_items.items())
    new_text = " | ".join([f"{name} x{qty}" for name, qty in new_items.items()])

    # 4. حساب المجموع النهائي (القديم + الجديد)
    # ملاحظة: لازم تكون مخزن عدد الأكياس القديمة في customer_data عند سحبها من الإكسيل
    old_qty = user.get("old_total_qty", 0) 
    final_qty = old_qty + new_qty
    final_products_price = old_products_price + new_price
    
    # 5. تقييم الشحن الجديد بناءً على إجمالي عدد الأكياس
    final_delivery = 0 if final_qty >= 5 else 30
    final_total_with_shipping = final_products_price + final_delivery
    
    combined_text = f"{old_order_text} + [إضافة: {new_text}]"
    
    # 6. التحديث في الإكسيل
    row_index = find_order_row_by_phone(user.get("temp_phone"))
    if row_index:
        delete_order_from_excel(row_index)
        delivery_status = "مجاني" if final_delivery == 0 else f"{final_delivery}ج"
        gift = "🎁 كيس هدية" if final_qty >= 8 else "لا يوجد"
        
        save_order_to_excel(data, combined_text, final_total_with_shipping, delivery_status, gift)
        
    return combined_text, final_total_with_shipping, final_delivery, final_products_price

def confirm_order(sender_id):
    user = USER_ORDERS.get(sender_id)
    if not user: return
    
    # --- الحالة الأولى: إضافة لطلب قائم ---
    if user.get("stage") == "adding_to_existing":
        # تم تصحيح الإزاحة هنا (كانت هناك مسافات زائدة)
        combined_text, combined_price, _, _ = update_existing_order_with_new_items(sender_id)
        
        # إشعار لبوت المتابعة بصيغة التعديل
        tracking_text = (
            "🔄 **تعديل طلب قائم (إضافة منتجات)**\n\n"
            f"👤 العميل: {user['customer_data'].get('الاسم ثلاثي')}\n"
            f"📞 الهاتف: {user.get('temp_phone')}\n"
            f"📝 الطلب الكامل بعد الإضافة: {combined_text}\n"
            f"💰 الإجمالي الجديد: {combined_price}ج"
        )
        # تأكد من أن هذه الدالة معرفة بهذا الاسم لديك
        send_telegram_notification(tracking_text)
        
        text = "🎉 تم تحديث طلبك بنجاح بإضافة المنتجات الجديدة!\nسيصلك الأوردر كاملاً في الموعد المحدد 🚚💚"
        send_message(sender_id, text)

    # --- الحالة الثانية: طلب جديد تماماً ---
    else:
        order = user.get("items", {})
        total_qty = sum(order.values())
        items_price = sum(PRODUCTS[name]*qty for name, qty in order.items())
        delivery_cost = 0 if total_qty >= 5 else 30
        total_price = items_price + delivery_cost
        delivery_text = "مجاني" if delivery_cost == 0 else f"{delivery_cost}ج"
        gift = "🎁 كيس هدية" if total_qty >= 8 else "لا يوجد"

        # 1. حساب رقم الأوردر
        try:
            from openpyxl import load_workbook
            wb = load_workbook("orders.xlsx")
            ws = wb.active
            order_number = ws.max_row
        except:
            order_number = 1

        # 2. تحويل القاموس لنص للإكسيل وحفظه
        excel_order_details = " | ".join([f"{name} x{qty}" for name, qty in order.items()])
        save_order_to_excel(user["customer_data"], excel_order_details, total_price, delivery_text, gift)

        # 3. تجهيز تفاصيل الطلب (الأصناف)
        details = "\n".join([f"✨ {name} x{qty} = {PRODUCTS[name]*qty}ج" for name, qty in order.items()])
        
        # 4. رسالة التليجرام بالصيغة الكاملة
        telegram_text = (
            f"🛒 **طلب جديد! رقم ({order_number})**\n\n"
            "👤 **بيانات العميل:**\n"
            f"الاسم ثلاثي: {user['customer_data'].get('الاسم ثلاثي','')}\n"
            f"اسم المحافظة: {user['customer_data'].get('اسم المحافظة','')}\n"
            f"اسم المنطقة: {user['customer_data'].get('اسم المنطقة','')}\n"
            f"اسم الشارع + علامة مميزة: {user['customer_data'].get('اسم الشارع + علامة مميزة','')}\n"
            f"رقم العمارة: {user['customer_data'].get('رقم العمارة','')}\n"
            f"رقم الشقة: {user['customer_data'].get('رقم الشقة','')}\n"
            f"رقم هاتف: {user['customer_data'].get('رقم هاتف ويفضل يكون عليه واتساب','')}\n"
            f"رقم هاتف آخر: {user['customer_data'].get('رقم هاتف اخر (ان وجد)','')}\n\n"
            "📦 **تفاصيل الطلب:**\n"
            f"{details}\n\n"
            f"💰 **الإجمالي:** {total_price}ج\n"
            f"🚚 **التوصيل:** {delivery_text}"
        )
        
        send_telegram_notification(telegram_text)

        # رسالة تأكيد للعميل على الفيس بوك
        special_area = user["customer_data"].get("اسم المنطقة","")
        if special_area in ["حلوان","15 مايو"]:
            text = "🎉 تم تأكيد طلب حضرتك بنجاح!\nطلبك هيوصل حضرتك يوم الثلاثاء القادم 🚚💚"
        else:
            text = "🎉 تم تأكيد طلب حضرتك بنجاح!\nطلبك هيوصل حضرتك في خلال 48 ساعة 🚚💚"
        send_message(sender_id, text)

    # --- تصفير الحالة والعودة للرئيسية ---
    USER_ORDERS[sender_id] = {
        "items": {},
        "data_fields": user.get("data_fields", []),
        "current_question": 0,
        "customer_data": {},
        "stage": "welcome"
    }
    send_main_menu(sender_id)
def cancel_order(sender_id):
    send_message(sender_id, "تم إلغاء الطلب بنجاح ❌")
    USER_ORDERS[sender_id]["items"] = {}
    USER_ORDERS[sender_id]["stage"] = "welcome"
    send_welcome(sender_id)

# ===== Run Flask =====
if __name__ == "__main__":
    app.run()
    app.run(port=5000, debug=True)